import openpyxl
from openpyxl.styles import Font, Alignment, colors, PatternFill
import sqlite3
import os
import pandas as pd


class SymbolDB:
    """
    Convert GoldWind Controller Symbol(Excel file) to an Isolated DataBase.
    """
    _symbol = None
    _db = None
    _tables = None
    _here = os.path.dirname(__file__)

    def __init__(self):
        pass

    def load_sym(self, sym_file):        
        if self.verification(sym_file, '.xlsx'):            
            self.src_file = sym_file        
            self.db_file = self.src_file.replace('.xlsx', '.db')
            self._symbol = openpyxl.load_workbook(self.src_file, read_only=True)
    
    def load_db(self, db_file):
        if self.verification(db_file, '.db'):
            self.db_file = db_file
            self.src_file = self.db_file.replace('.db', '.xlsx')

    def verification(self, file, post_fix):
        if os.path.isfile(file) and os.path.splitext(file)[1] == post_fix:
            return True
        else:
            raise FileNotFoundError(f"文件不存在或不是{post_fix}格式！")

    def connect(self):
        self._db = sqlite3.connect(self.db_file)
        self._tables = self.get_tables()

    def is_connect(self):
        if isinstance(self._db, sqlite3.Connection):
            return True

        return False

    def close(self):
        self._db.close()

    def create_db(self, omit_exist=True):
        """
        Will delete existent .db file
        """        
        
        if not os.path.exists(self.db_file):            
            self._create_db()
        elif os.path.exists(self.db_file) and not omit_exist:
            os.remove(self.db_file)
            self._create_db()
        else:
            pass
        if self.is_connect():
            self.close()

    def _create_db(self):
        self.connect()
        sheet_names = self._symbol.sheetnames
        for sheet in sheet_names:
            if sheet in ['Settings', 'State Machine']:
                continue
            self._create_table(sheet)

        self._db.commit()

    def _create_table(self, sheet):
        cursor = self._db.cursor()
        ws = self._symbol[sheet]
        iter_row = ws.rows
        column_names = {}
        table_name = sheet.replace(' ', '_')

        while True:
            try:
                row = next(iter_row)
            except StopIteration:
                break

            if not row:  # 过滤空行
                continue

            start_with = row[0].value

            row_values = [c.value for c in row]
            
            if start_with and ';' in start_with:
                continue            
            elif start_with in ['Name', 'Channel Name']:
                column_names = self.__get_column_names(row_values)
                column_name_str = self.__get_column_name_str(column_names)
                unique_item = self.__get_unique_item(sheet, start_with)
                sql_statement = f"CREATE TABLE {table_name} (" \
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, " \
                    f"{column_name_str}, " \
                    f"CONSTRAINT name_unique UNIQUE ({unique_item}))"
                
                cursor.execute(sql_statement)

                # 给表加索引有时会失败， 此时忽略
                try:
                    cursor.execute(f"CREATE UNIQUE INDEX index_name ON {table_name} ({start_with.replace(' ', '_')})")
                except sqlite3.OperationalError:
                    pass
            else:
                column_value_str = self.__get_column_value_str(column_names, row_values)
                column_name_str = self.__get_column_name_str(column_names, no_text=True)
                sql_statement = f"INSERT INTO {table_name} ({column_name_str}) VALUES ({column_value_str})"
                try:
                    cursor.execute(sql_statement)
                except:
                    pass

        cursor.close()

    @staticmethod
    def __get_column_names(row_values):
        column_names = {f"{str(n).replace(' ', '_')}": i for i, n in enumerate(row_values) 
                        if n is not None and ';' not in str(n)}

        return column_names

    @staticmethod
    def __get_column_name_str(column_names, no_text=False):
        if no_text:
            column_name_str = ', '.join([f"_{str(n)}" if str(n).isnumeric() else f"{str(n)}" 
                          for n in column_names.keys()])
        else:
            column_name_str = ', '.join([f"_{str(n)} text" if str(n).isnumeric() else f"{str(n)} text" 
                          for n in column_names.keys()])

        return column_name_str

    @staticmethod
    def __get_unique_item(sheet, start_with):
        unique_item = "id" if sheet in ['Schedules', 'Filters'] else f"id, {start_with.replace(' ', '_')}"

        return unique_item

    @staticmethod
    def __get_column_value_str(column_names, row_values):
        """数据库中的值全部为字符串格式"""
        column_value_str = ', '.join([f"'{str(row_values[i])}'" for i in column_names.values()])

        return column_value_str

    @property
    def tables(self):
        return self._tables

    def get_tables(self):
        """ 
        return all tables in self._db
        """
        cursor = self._db.cursor()
        sql_statement = 'select name from sqlite_master where type="table"'
        cursor.execute(sql_statement)
        result = cursor.fetchall()
        cursor.close()
        tables = [r[0] for r in result]

        return tables

    def query(self, keyword, sheet_name=None):
        if not self.is_connect():
            return []

        if sheet_name in ['Schedules', 'Filters']:
            # T表返回DataFrame格式, 存放于以变量名为键值的字典中
            table_name = sheet_name.replace(' ', '_')
            sql_statement = f"SELECT Name FROM {table_name} WHERE Name LIKE '%{keyword}%'"
            cursor = self._db.cursor()
            cursor.execute(sql_statement)
            all_name = [n[0] for n in cursor.fetchall()]

            result = {k: v for k, v in self._TData_result(table_name).items() if k in all_name}
            return result

        if sheet_name:
            table_name = sheet_name.replace(' ', '_')
            key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name' 
            sql_statement = f"SELECT * FROM {table_name} WHERE {key_name} LIKE '%{keyword}%'"
        else:
            all_tables = self.tables
            sql_statement_list = []
            for table_name in all_tables:
                key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name' 
                sql_statement_list.append(f"SELECT {key_name} FROM {table_name} WHERE {key_name} LIKE '%{keyword}%'")
            sql_statement = ' UNION '.join(sql_statement_list)            

        cursor = self._db.cursor()
        cursor.execute(sql_statement)
        result = cursor.fetchall()

        result_no_id = [r[1:] for r in result] if sheet_name else [r[0] for r in result]

        return  result_no_id

    def _TData_result(self, table_name):
        """Filter and Schedule Data"""
        sql_statement = f"SELECT * FROM {table_name}"

        columns = self._db.execute(f'pragma table_info({table_name})').fetchall()
        cloumns = [c[1] for c in columns]
        
        # execute shortcut 
        result = self._db.execute(sql_statement).fetchall()
        df_dict = self._data_classify(result, cloumns)

        return df_dict
        
    @staticmethod
    def _data_classify(data, tables):
        _dict = {}
        for row in data:
            if row[1] != 'None':
                key_name = row[1]
                _dict[key_name] = [row[3:]]
            else:
                _dict[key_name].append(row[3:])
        df_dict = {}
        for k, v in _dict.items():
            df_dict[k] = pd.DataFrame(v, columns=tables[3:])

        return df_dict

    def update(self, **kwargs):
        """
        kwargs =
        {
            name1: {
                sheet: "",
                value: ""
            },            
            name2: {
                sheet: "",
                value: ""
            }
        }
        """
        if not self.verification(self.src_file, '.xlsx'):
            return None
        wb = openpyxl.load_workbook(self.src_file)
        sheet_names_to_update = set([v['sheet'] for k, v in kwargs.items()])
        sheets = [wb[sheet_name] for sheet_name in sheet_names_to_update]
        
        for sheet in sheets:
            iter_row = sheet.rows
            irow = 0

            while True:
                try:
                    row = next(iter_row)
                    irow += 1
                except StopIteration:
                    break

                if not row:  # 过滤空行
                    continue

                start_with = row[0].value
                row_values = [c.value for c in row]
                column_names = []
                if start_with in ['Name', 'Channel Name']:
                    column_names = self.__get_column_names(row_values)
                elif start_with in kwargs.keys():
                    if 'P_' in start_with:
                        icol = column_names.index('Initial Value') + 2
                        sheet.cell(irow, icol, kwargs[start_with]['value'])
                        font = sheet.cell(irow, icol).font
                        alignment = sheet.cell(irow, icol).alignment
                        fill = PatternFill(patternType=sheet.cell(irow, icol).fill.patternType,
                                           bgColor=sheet.cell(irow, icol).fill.bgColor,
                                           fgColor=sheet.cell(irow, icol).fill.fgColor)
                        
                        sheet.cell(irow, icol).font = font
                        sheet.cell(irow, icol).alignment = alignment
                        sheet.cell(irow, icol).fil = fill
                        
                        self._update_P(sheet, start_with, kwargs)
                    if 'T_' in start_with:
                        self._update_T(sheet, start_with, kwargs)
                    if 'F_' in start_with:
                        self._update_F(sheet, start_with, kwargs)

    # def _update_P(self, sheet, start_with, kwargs):
