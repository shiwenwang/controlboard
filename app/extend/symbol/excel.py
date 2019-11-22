import openpyxl
from openpyxl.cell.read_only import EmptyCell
import sqlite3
import os
import pandas as pd
import re


class SymbolDB:
    """
    Convert GoldWind Controller Symbol(Excel file) to an Isolated DataBase.
    """
    _symbol = None
    _db = None
    _tables = None
    _here = os.path.dirname(__file__)

    def __init__(self):
        pass

    def get_wb(self):
        return self._symbol

    def load_sym(self, sym_file, db_name=None):
        if self.verification(sym_file, '.xlsx'):
            self.src_file = sym_file
            self.db_file = self.src_file.replace('.xlsx', '.db') if db_name is None else os.path.join(
                os.path.split(self.src_file)[0], db_name + '.db')
            self._symbol = openpyxl.load_workbook(
                self.src_file, read_only=True)

    def load_db(self, db_file, excel_name=None):
        if self.verification(db_file, '.db'):
            self.db_file = db_file
            self.src_file = self.db_file.replace('.db', '.xlsx') if excel_name is None else os.path.join(
                os.path.split(self.db_file)[0], excel_name)

    def verification(self, file, post_fix):
        if os.path.isfile(file) and os.path.splitext(file)[1] == post_fix:
            return True
        else:
            raise FileNotFoundError(f"文件不存在或不是{post_fix}格式！")

    def connect(self):
        self._db = sqlite3.connect(self.db_file)
        self._tables = self.get_tables()

    def is_connect(self):
        if isinstance(self._db, sqlite3.Connection):
            return True

        return False

    def close(self):
        self._db.close()

    def create_db(self, omit_exist=True):
        """
        Will delete existent .db file
        """

        if not os.path.exists(self.db_file):
            self._create_db()
        elif os.path.exists(self.db_file) and not omit_exist:
            os.remove(self.db_file)
            self._create_db()
        else:
            pass
        if self.is_connect():
            self.close()

        self._symbol.close()

    def _create_db(self):
        self.connect()
        sheet_names = self._symbol.sheetnames
        for sheet in sheet_names:
            if sheet in ['Settings', 'State Machine']:
                continue
            self._create_table(sheet, creat_pos=True)

        self._db.commit()

    def _create_table(self, sheet, creat_pos=False):
        cursor = self._db.cursor()
        ws = self._symbol[sheet]
        iter_row = ws.rows
        column_names = {}
        table_name = sheet.replace(' ', '_')

        n_row = 0
        while True:
            try:
                row = next(iter_row)
                n_row = n_row + 1
            except StopIteration:
                break

            if not row or all([isinstance(r, EmptyCell) for r in row]):  # 过滤空行
                continue

            start_with = row[0].value
            if start_with and ';' in start_with:
                continue

            row_values = [c.value for c in row]

            if row_values[0] is not None:
                last_name = row_values[0]
            else:
                row_values[0] = last_name

            if creat_pos:
                positions = [(n_row, i + 1) for i, value in enumerate(row_values)]

            if start_with in ['Name', 'Channel Name']:
                column_names = self.__get_column_names(row_values)
                column_name_str = self.__get_column_name_str(column_names)
                unique_item = self.__get_unique_item(sheet, start_with)

                # 创建数值表
                sql_statement = f"CREATE TABLE {table_name} (" \
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, " \
                    f"{column_name_str}, " \
                    f"CONSTRAINT name_unique UNIQUE ({unique_item}))"
                cursor.execute(sql_statement)

                # 创建位置表
                if creat_pos:
                    sql_statement = f"CREATE TABLE {table_name}_pos (" \
                        "id INTEGER PRIMARY KEY AUTOINCREMENT, " \
                        f"{column_name_str}, " \
                        f"CONSTRAINT name_unique UNIQUE ({unique_item}))"
                    cursor.execute(sql_statement)

                # 给表加索引有时会失败， 此时忽略
                try:
                    cursor.execute(
                        f"CREATE UNIQUE INDEX index_name ON {table_name} ({start_with.replace(' ', '_')})")
                    if creat_pos:
                        cursor.execute(
                            f"CREATE UNIQUE INDEX index_name ON {table_name}_pos ({start_with.replace(' ', '_')})")
                except sqlite3.OperationalError:
                    pass
            else:
                column_value_str = self.__get_column_value_str(
                    column_names, row_values)
                column_name_str = self.__get_column_name_str(
                    column_names, no_text=True)
                sql_statement = f"INSERT INTO {table_name} ({column_name_str}) VALUES ({column_value_str})"
                cursor.execute(sql_statement)

                if creat_pos:
                    positions[0] = row_values[0]
                    positions_str = self.__get_value_position_str(column_names, positions)
                    sql_statement = f"INSERT INTO {table_name}_pos ({column_name_str}) VALUES ({positions_str})"
                    cursor.execute(sql_statement)

        cursor.close()

    @staticmethod
    def __get_column_names(row_values):
        column_names = {f"{str(n).replace(' ', '_')}": i for i, n in enumerate(row_values)
                        if n is not None and ';' not in str(n)}

        return column_names

    @staticmethod
    def __get_column_name_str(column_names, no_text=False):
        if no_text:
            column_name_str = ', '.join([f"_{str(n)}" if str(n).isnumeric() else f"{str(n)}"
                                         for n in column_names.keys()])
        else:
            column_name_str = ', '.join([f"_{str(n)} text" if str(n).isnumeric() else f"{str(n)} text"
                                         for n in column_names.keys()])

        return column_name_str

    @staticmethod
    def __get_unique_item(sheet, start_with):
        # unique_item = "id" if sheet in [
        #     'Schedules', 'Filters'] else f"id, {start_with.replace(' ', '_')}"
        unique_item = 'id'

        return unique_item

    @staticmethod
    def __get_column_value_str(column_names, row_values):
        """数据库中的值全部为字符串格式"""
        column_value_str = ', '.join(
            [f"'{str(row_values[i])}'" for i in column_names.values()])

        return column_value_str

    @staticmethod
    def __get_value_position_str(column_names, positions):
        """数据库中的值全部为字符串格式"""
        column_value_str = ', '.join(
            [f"'{str(positions[i])}'" for i in column_names.values()])

        return column_value_str

    @property
    def tables(self):
        return self._tables

    def get_tables(self):
        """ 
        return all tables in self._db
        """
        cursor = self._db.cursor()
        sql_statement = 'select name from sqlite_master where type="table"'
        cursor.execute(sql_statement)
        result = cursor.fetchall()
        cursor.close()
        tables = [r[0] for r in result if '_pos' not in r[0]]

        return tables

    def belong_to(self, param):
        cursor = self._db.cursor()
        for table in self.tables:
            key_name = 'Channel_Name' if table == 'DISCON_Mappings' else 'Name'
            cursor.execute(f"SELECT * FROM {table} WHERE {key_name}=?", (param,))
            if cursor.fetchone() is not None:
                return table

    def multi_query(self, keywords):
        if not isinstance(keywords, list):
            return self.query(keywords, self.belong_to(keywords))
        # 隶属关系
        relationship = {}
        for keyword in keywords:
            table_name = self.belong_to(keyword)
            if table_name is None:
                continue
            if table_name not in relationship.keys():
                relationship.update({table_name: [keyword]})
            relationship[table_name].append(keyword)

        result = {}
        for k, v in relationship.items():
            result.update(self._multi_query(v, k))

        return result

    def _multi_query(self, keywords, table_name):
        key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name'
        cursor = self._db.cursor()
        sql_statement = f"SELECT * FROM {table_name} WHERE {key_name} in {tuple(keywords)}"
        cursor.execute(sql_statement)
        values = cursor.fetchall()

        cursor.execute(f"PRAGMA table_info({table_name})")
        keys_info = cursor.fetchall()
        cursor.close()

        keys = [k[1] for k in keys_info]

        result = {}
        for value in values:
            # table变量名追加ID
            key_name = value[1] + str(value[0]) if value[1][:2] in ['T_', 'F_'] else value[1]
            result.update({key_name: pd.Series(data=value, index=keys)})

        return result

    def query(self, keyword, sheet_name=None):
        """
        sheet_name = None： 将只返回Name列表
        sheet_name ！= None： 返回所有字段
        :param keyword:
        :param sheet_name:
        :return:
        """
        if not self.is_connect():
            return []

        if sheet_name in ['Schedules', 'Filters']:
            # T表返回DataFrame格式, 存放于以变量名为键值的字典中
            table_name = sheet_name.replace(' ', '_')
            sql_statement = f"SELECT Name FROM {table_name} WHERE Name LIKE '%{keyword}%'"
            cursor = self._db.cursor()
            cursor.execute(sql_statement)
            all_name = [n[0] for n in cursor.fetchall()]

            result = {k: v for k, v in self._TData_result(
                table_name).items() if k in all_name}
            return result

        if sheet_name is None:
            all_tables = ['State_Filter', 'Alarm', 'Auto_System', 'CIO', 'State', 'Derived', 'Param',
                          'Filters', 'Schedules', 'DISCON_Parameters', 'DISCON_Mappings']
            sql_statement_list = []
            for table_name in all_tables:
                key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name'
                sql_statement_list.append(
                    f"SELECT {key_name} FROM {table_name} WHERE {key_name} LIKE '%{keyword}%' escape '/'")
            sql_statement = ' UNION '.join(sql_statement_list)

            cursor = self._db.cursor()
            cursor.execute(sql_statement)
            values = cursor.fetchall()
            result = [r for r in values]
            return result

        table_name = sheet_name.replace(' ', '_')
        key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name'
        sql_statement = f"SELECT * FROM {table_name} WHERE {key_name} LIKE '%{keyword}%'"

        cursor = self._db.cursor()
        cursor.execute(sql_statement)
        values = cursor.fetchall()

        cursor.execute(f"PRAGMA table_info({table_name})")
        keys_info = cursor.fetchall()
        cursor.close()

        keys = [k[1] for k in keys_info]

        result = {v[1]: pd.Series(data=v, index=keys) for v in values}

        return result

    def _TData_result(self, table_name):
        """Filter and Schedule Data"""
        sql_statement = f"SELECT * FROM {table_name}"

        columns = self._db.execute(
            f'pragma table_info({table_name})').fetchall()
        cloumns = [c[1] for c in columns]

        # execute shortcut
        result = self._db.execute(sql_statement).fetchall()
        df_dict = self._data_classify(result, cloumns)

        return df_dict

    @staticmethod
    def _data_classify(data, tables):
        _dict = {}
        for row in data:
            if row[1] != 'None':
                key_name = row[1]
                _dict[key_name] = [row[3:]]
            else:
                _dict[key_name].append(row[3:])
        df_dict = {}
        for k, v in _dict.items():
            df_dict[k] = pd.DataFrame(v, columns=tables[3:])

        return df_dict

    def update(self, target, **kwargs):
        # 变量和表的隶属关系
        relationship = {}
        for keyword in kwargs.keys():
            pattern = re.compile(r'\d+-\S*$')
            fine_keyword = keyword.split('-')[0] if 'P_' in keyword else pattern.sub('', keyword, 1)
            table_name = self.belong_to(fine_keyword)
            if table_name is None:
                continue
            if table_name not in relationship.keys():
                relationship.update({table_name: [keyword]})
            else:
                relationship[table_name].append(keyword)

        self.db_update(relationship, kwargs)
        if target == 'symbol':
            self.excel_update(relationship, kwargs)

    def excel_update(self, relationship, kwargs):
        """
        更新Symbol表
        """
        from time import time
        start = time()
        wb = openpyxl.load_workbook(self.src_file)
        print(f'读： {time() - start}')
        sheets = [wb[table_name.replace('_', ' ')] for table_name in relationship.keys()]
        for table_name, params in relationship.items():
            sheet_name = table_name.replace('_', ' ')
            ws = wb[sheet_name]
            key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name'
            cursor = self._db.cursor()
            if table_name in ['Filters', 'Schedules']:
                pass
            else:
                for param in params:
                    sql_statement = f'SELECT Initial_Value FROM {table_name}_pos WHERE {key_name} = "{param.split("-")[0]}"'
                    cursor.execute(sql_statement)
                    result = cursor.fetchall()[0][0]
                    row, col = eval(result)
                    ws.cell(column=col, row=row, value=kwargs[param])
        print(f'写： {time() - start}')
        wb.save(self.src_file)
        print(f'存： {time() - start}')

    def db_update(self, relationship, kwargs):
        """
        更新数据库
        """
        cursor = self._db.cursor()
        for table_name, params in relationship.items():
            # 更新表
            key_name = 'Channel_Name' if table_name == 'DISCON_Mappings' else 'Name'
            if table_name in ['Filters', 'Schedules']:
                # 逐条记录更新
                pattern = re.compile(r'(\d+)-(\S+)$')
                set_statement = {}
                for param in params:
                    m = pattern.search(param)
                    id, col = m.groups()
                    col = f'_{col}' if col.isdigit() else col
                    set_statement[id] = set_statement[id] + f', {col} = "{kwargs[param]}"' \
                        if id in set_statement.keys() else f'{col} = "{kwargs[param]}"'

                for id, sta in set_statement.items():
                    sql_statement = f'UPDATE {table_name} SET {sta} WHERE id = {id}'
                    cursor.execute(sql_statement)
            else:
                # 一次更新
                case_statement = []
                true_params = [p.split('-')[0] for p in params]
                for param in params:
                    case_statement.append(f'WHEN "{param.split("-")[0]}" THEN "{kwargs[param]}"')

                sql_statement = f'UPDATE {table_name} SET Initial_Value = CASE {key_name} ' \
                    f'{" ".join(case_statement)}' \
                    f' END WHERE {key_name} IN {tuple(true_params)}' \
                    if len(params) > 1 else \
                    f'UPDATE {table_name} SET Initial_Value = "{list(kwargs.values())[0]}" WHERE {key_name} ' \
                        f'= "{true_params[0]}"'

                cursor.execute(sql_statement)
        cursor.close()
        self._db.commit()
